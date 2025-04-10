import os
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from module import parse_zones, parse_user_groups, parse_users, parse_address_groups, parse_hosts, parse_dhcp_pools, parse_interfaces, parse_system_info, parse_ip_pools, parse_virtual_ips, parse_firewall_policies, parse_routes, parse_route_policies, parse_ipsec

def create_excel_report(system_info, interfaces, dhcpv4_pools, dhcpv6_pools, hostsv4, hostsv6, address_groups, admins, users, user_groups, zones, ip_pools, virtual_ips, firewall_policies_v4, firewall_policies_v6, routes, route_policies, ipsec_configs, input_file):
    # Create Excel file
    wb = openpyxl.Workbook()

    ### Create System tab
    ws_system = wb.active
    ws_system.title = "Systèmes"

    # Create headers
    headers = ["Hostname", "Modèle", "Version", "Build"]
    for col, header in enumerate(headers, 1):
        cell = ws_system.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill system data
    for col, header in enumerate(headers, 1):
        ws_system.cell(row=2, column=col).value = system_info.get(header, "")

    # Adjust column width
    for col in ws_system.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_system.column_dimensions[column].width = adjusted_width

    ### Create Interfaces tab
    ws_interfaces = wb.create_sheet(title="Interfaces")

    # Create headers, now including IPv6 columns
    interface_headers = [
        "Interface", "Alias", "Vdom", "Adresse IPv4", "NetMask", "Adresse IPv6", "VLAN ID", "Ip Secondaire",
        "Accès IPv4", "Accès IPv6", "Zone", "Rôle", "Type", "Membre", "Tag", "Description"
    ]

    for col, header in enumerate(interface_headers, 1):
        cell = ws_interfaces.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill interface data with alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row, interface in enumerate(interfaces, 2):
        # Apply alternating colors
        if row % 2 == 0:  # Even rows
            for col in range(1, len(interface_headers) + 1):
                ws_interfaces.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(interface_headers, 1):
            ws_interfaces.cell(row=row, column=col).value = interface.get(header, "")

    # Adjust column width
    for col in ws_interfaces.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_interfaces.column_dimensions[column].width = adjusted_width

    # Add filters to headers
    last_col_letter = openpyxl.utils.get_column_letter(len(interface_headers))
    ws_interfaces.auto_filter.ref = f"A1:{last_col_letter}{len(interfaces) + 1}"

    ### Create DHCPV4 Pool tab
    ws_dhcpv4 = wb.create_sheet(title="DHCP Pool IPv4")

    # Create headers for DHCP Pool tab
    dhcpv4_headers = ["Interface", "Domaine", "Gateway", "NetMask", "Start Pool", "End Pool", "DNS"]

    for col, header in enumerate(dhcpv4_headers, 1):
        cell = ws_dhcpv4.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill DHCP Pool data with alternating row colors
    for row, pool in enumerate(dhcpv4_pools, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(dhcpv4_headers) + 1):
                ws_dhcpv4.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(dhcpv4_headers, 1):
            ws_dhcpv4.cell(row=row, column=col).value = pool.get(header, "")

    # Adjust column width for DHCP Pool tab
    for col in ws_dhcpv4.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_dhcpv4.column_dimensions[column].width = adjusted_width

    # Add filters to DHCP Pool tab headers
    last_col_letter_dhcp = openpyxl.utils.get_column_letter(len(dhcpv4_headers))
    ws_dhcpv4.auto_filter.ref = f"A1:{last_col_letter_dhcp}{len(dhcpv4_pools) + 1}"

    ### Create DHCPV6 Pool tab
    ws_dhcpv6 = wb.create_sheet(title="DHCP Pool IPv6")

    # Create headers for DHCP Pool IPv6 tab
    dhcpv6_headers = ["Interface", "Network", "Start Pool", "End Pool", "DNS"]

    for col, header in enumerate(dhcpv6_headers, 1):
        cell = ws_dhcpv6.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill DHCP Pool IPv6 data with alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row, pool in enumerate(dhcpv6_pools, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(dhcpv6_headers) + 1):
                ws_dhcpv6.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(dhcpv6_headers, 1):
            ws_dhcpv6.cell(row=row, column=col).value = pool.get(header, "")

    # Adjust column width for DHCP Pool IPv6 tab
    for col in ws_dhcpv6.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_dhcpv6.column_dimensions[column].width = adjusted_width

    # Add filters to DHCP Pool IPv6 tab headers
    last_col_letter_dhcpv6 = openpyxl.utils.get_column_letter(len(dhcpv6_headers))
    ws_dhcpv6.auto_filter.ref = f"A1:{last_col_letter_dhcpv6}{len(dhcpv6_pools) + 1}"

    ### Create Host tab
    ws_hostsv4 = wb.create_sheet(title="Host IPv4")

    # Create headers for Host tab
    hostv4_headers = ["Hostname", "UUID", "Interface", "Type", "MAC", "Adresse", "NetMask", "FQDN", "Start IP", "End IP", "Commentaire"]

    for col, header in enumerate(hostv4_headers, 1):
        cell = ws_hostsv4.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Host data with alternating row colors
    for row, host in enumerate(hostsv4, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(hostv4_headers) + 1):
                ws_hostsv4.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(hostv4_headers, 1):
            ws_hostsv4.cell(row=row, column=col).value = host.get(header, "")

    # Adjust column width for Host tab
    for col in ws_hostsv4.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_hostsv4.column_dimensions[column].width = adjusted_width

    # Add filters to Host tab headers
    last_col_letter_host = openpyxl.utils.get_column_letter(len(hostv4_headers))
    ws_hostsv4.auto_filter.ref = f"A1:{last_col_letter_host}{len(hostsv4) + 1}"

    ### Create Host IPv6 tab
    ws_hostsv6 = wb.create_sheet(title="Host IPv6")

    # Create headers for Host IPv6 tab
    hostsv6_headers = ["Hostname", "UUID", "Adresse IPv6"]

    for col, header in enumerate(hostsv6_headers, 1):
        cell = ws_hostsv6.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Host IPv6 data with alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row, host in enumerate(hostsv6, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(hostsv6_headers) + 1):
                ws_hostsv6.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(hostsv6_headers, 1):
            ws_hostsv6.cell(row=row, column=col).value = host.get(header, "")

    # Adjust column width for Host IPv6 tab
    for col in ws_hostsv6.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_hostsv6.column_dimensions[column].width = adjusted_width

    # Add filters to Host IPv6 tab headers
    last_col_letter_hostsv6 = openpyxl.utils.get_column_letter(len(hostsv6_headers))
    ws_hostsv6.auto_filter.ref = f"A1:{last_col_letter_hostsv6}{len(hostsv6) + 1}"

    ### Create Host Groupe tab
    ws_host_groupe = wb.create_sheet(title="Host Groupe")

    # Create headers for Host Groupe tab
    host_groupe_headers = ["Groupe", "UUID", "Membre"]

    for col, header in enumerate(host_groupe_headers, 1):
        cell = ws_host_groupe.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Host Groupe data with alternating row colors
    for row, group in enumerate(address_groups, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(host_groupe_headers) + 1):
                ws_host_groupe.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(host_groupe_headers, 1):
            ws_host_groupe.cell(row=row, column=col).value = group.get(header, "")

    # Adjust column width for Host Groupe tab
    for col in ws_host_groupe.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_host_groupe.column_dimensions[column].width = adjusted_width

    # Add filters to Host Groupe tab headers
    last_col_letter_groupe = openpyxl.utils.get_column_letter(len(host_groupe_headers))
    ws_host_groupe.auto_filter.ref = f"A1:{last_col_letter_groupe}{len(address_groups) + 1}"

 ### Create Admin tab
    ws_admin = wb.create_sheet(title="Admin")

    # Create headers for Admin tab
    admin_headers = ["Name", "Profil Type", "Vdom", "Trust Host IPv4", "Trust Host IPv6"]

    for col, header in enumerate(admin_headers, 1):
        cell = ws_admin.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Admin data with alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for row, admin in enumerate(admins, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(admin_headers) + 1):
                ws_admin.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(admin_headers, 1):
            ws_admin.cell(row=row, column=col).value = admin.get(header, "")

    # Adjust column width for Admin tab
    for col in ws_admin.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_admin.column_dimensions[column].width = adjusted_width

    # Add filters to Admin tab headers
    last_col_letter_admin = openpyxl.utils.get_column_letter(len(admin_headers))
    ws_admin.auto_filter.ref = f"A1:{last_col_letter_admin}{len(admins) + 1}"

    ### Create User tab
    ws_users = wb.create_sheet(title="User")

    # Create headers for User tab
    user_headers = ["User", "Type", "MFA", "Mail"]

    for col, header in enumerate(user_headers, 1):
        cell = ws_users.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill User data with alternating row colors
    for row, user in enumerate(users, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(user_headers) + 1):
                ws_users.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(user_headers, 1):
            ws_users.cell(row=row, column=col).value = user.get(header, "")

    # Adjust column width for User tab
    for col in ws_users.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_users.column_dimensions[column].width = adjusted_width

    # Add filters to User tab headers
    last_col_letter_user = openpyxl.utils.get_column_letter(len(user_headers))
    ws_users.auto_filter.ref = f"A1:{last_col_letter_user}{len(users) + 1}"

    ### Create Groupe tab
    ws_groups = wb.create_sheet(title="Groupe")

    # Create headers for Groupe tab
    group_headers = ["Groupe", "Membre"]

    for col, header in enumerate(group_headers, 1):
        cell = ws_groups.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Groupe data with alternating row colors
    for row, group in enumerate(user_groups, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(group_headers) + 1):
                ws_groups.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(group_headers, 1):
            ws_groups.cell(row=row, column=col).value = group.get(header, "")

    # Adjust column width for Groupe tab
    for col in ws_groups.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_groups.column_dimensions[column].width = adjusted_width

    # Add filters to Groupe tab headers
    last_col_letter_group = openpyxl.utils.get_column_letter(len(group_headers))
    ws_groups.auto_filter.ref = f"A1:{last_col_letter_group}{len(user_groups) + 1}"

    ### Create Zone tab
    ws_zones = wb.create_sheet(title="Zone")

    # Create headers for Zone tab
    zone_headers = ["Zone", "Membre"]

    for col, header in enumerate(zone_headers, 1):
        cell = ws_zones.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Zone data with alternating row colors
    for row, zone in enumerate(zones, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(zone_headers) + 1):
                ws_zones.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(zone_headers, 1):
            ws_zones.cell(row=row, column=col).value = zone.get(header, "")

    # Adjust column width for Zone tab
    for col in ws_zones.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_zones.column_dimensions[column].width = adjusted_width

    # Add filters to Zone tab headers
    last_col_letter_zone = openpyxl.utils.get_column_letter(len(zone_headers))
    ws_zones.auto_filter.ref = f"A1:{last_col_letter_zone}{len(zones) + 1}"

    ### Create Ip Pool tab
    ws_ip_pools = wb.create_sheet(title="Ip Pool")

    # Create headers for Ip Pool tab
    ip_pool_headers = ["Name", "Type", "Start IP", "End IP", "Source Start IP", "Source End IP", "ARP"]

    for col, header in enumerate(ip_pool_headers, 1):
        cell = ws_ip_pools.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Ip Pool data with alternating row colors
    for row, pool in enumerate(ip_pools, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(ip_pool_headers) + 1):
                ws_ip_pools.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(ip_pool_headers, 1):
            ws_ip_pools.cell(row=row, column=col).value = pool.get(header, "")

    # Adjust column width for Ip Pool tab
    for col in ws_ip_pools.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_ip_pools.column_dimensions[column].width = adjusted_width

    # Add filters to Ip Pool tab headers
    last_col_letter_ip_pool = openpyxl.utils.get_column_letter(len(ip_pool_headers))
    ws_ip_pools.auto_filter.ref = f"A1:{last_col_letter_ip_pool}{len(ip_pools) + 1}"

    ### Create Virtual IP tab
    ws_virtual_ips = wb.create_sheet(title="Virtual IP")

    # Create headers for Virtual IP tab
    vip_headers = ["Name", "UUID", "External IP", "Map IP", "External Port", "Map Port", "Protocole", "External Interface", "Port Forward"]

    for col, header in enumerate(vip_headers, 1):
        cell = ws_virtual_ips.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Virtual IP data with alternating row colors
    for row, vip in enumerate(virtual_ips, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(vip_headers) + 1):
                ws_virtual_ips.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(vip_headers, 1):
            ws_virtual_ips.cell(row=row, column=col).value = vip.get(header, "")

    # Adjust column width for Virtual IP tab
    for col in ws_virtual_ips.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_virtual_ips.column_dimensions[column].width = adjusted_width

    # Add filters to Virtual IP tab headers
    last_col_letter_vip = openpyxl.utils.get_column_letter(len(vip_headers))
    ws_virtual_ips.auto_filter.ref = f"A1:{last_col_letter_vip}{len(virtual_ips) + 1}"

    ### Create Règle tab v4
    ws_regles = wb.create_sheet(title="Règle IPv4")

    # Create headers for Règle tab
    regle_headers = ["Name", "UUID", "Source Interface", "Destination Interface", "Source Adresse", "Destination Adresse",
                     "Service", "Planification", "Groupe", "User", "NAT", "IP Pool", "Pool Name", "Action", "Log Traffic", "Commentaire"]

    for col, header in enumerate(regle_headers, 1):
        cell = ws_regles.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Règle data with alternating row colors
    for row, regle in enumerate(firewall_policies_v4, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(regle_headers) + 1):
                ws_regles.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(regle_headers, 1):
            ws_regles.cell(row=row, column=col).value = regle.get(header, "")

    # Adjust column width for Règle tab
    for col in ws_regles.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_regles.column_dimensions[column].width = adjusted_width

    # Add filters to Règle tab headers
    last_col_letter_regle = openpyxl.utils.get_column_letter(len(regle_headers))
    ws_regles.auto_filter.ref = f"A1:{last_col_letter_regle}{len(firewall_policies_v4) + 1}"

    ### Create Règle tab v6
    ws_regles = wb.create_sheet(title="Règle IPv6")

    # Create headers for Règle tab
    regle_headers = ["Name", "UUID", "Source Interface", "Destination Interface", "Source Adresse", "Destination Adresse",
                     "Service", "Planification", "Groupe", "User", "NAT", "IP Pool", "Pool Name", "Action", "Log Traffic", "Commentaire"]

    for col, header in enumerate(regle_headers, 1):
        cell = ws_regles.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Règle data with alternating row colors
    for row, regle in enumerate(firewall_policies_v6, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(regle_headers) + 1):
                ws_regles.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(regle_headers, 1):
            ws_regles.cell(row=row, column=col).value = regle.get(header, "")

    # Adjust column width for Règle tab
    for col in ws_regles.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_regles.column_dimensions[column].width = adjusted_width

    # Add filters to Règle tab headers
    last_col_letter_regle = openpyxl.utils.get_column_letter(len(regle_headers))
    ws_regles.auto_filter.ref = f"A1:{last_col_letter_regle}{len(firewall_policies_v6) + 1}"

    ### Create Route tab
    ws_routes = wb.create_sheet(title="Route")

    # Create headers for Route tab
    route_headers = ["Destination", "Interface", "Gateway", "Priorité", "Distance Administrative", "Blackhole", "Commentaire"]

    for col, header in enumerate(route_headers, 1):
        cell = ws_routes.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Route data with alternating row colors
    for row, route in enumerate(routes, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(route_headers) + 1):
                ws_routes.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(route_headers, 1):
            ws_routes.cell(row=row, column=col).value = route.get(header, "")

    # Adjust column width for Route tab
    for col in ws_routes.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_routes.column_dimensions[column].width = adjusted_width

    # Add filters to Route tab headers
    last_col_letter_route = openpyxl.utils.get_column_letter(len(route_headers))
    ws_routes.auto_filter.ref = f"A1:{last_col_letter_route}{len(routes) + 1}"

    ### Create Route Policy tab
    ws_route_policies = wb.create_sheet(title="Route Policy")

    # Create headers for Route Policy tab
    route_policy_headers = ["Source Interface", "Destination Interface", "Source Adresse", "Destination Adresse",
                            "Start Port", "End Port", "Gateway", "Status", "Action"]

    for col, header in enumerate(route_policy_headers, 1):
        cell = ws_route_policies.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill Route Policy data with alternating row colors
    for row, policy in enumerate(route_policies, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(route_policy_headers) + 1):
                ws_route_policies.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(route_policy_headers, 1):
            ws_route_policies.cell(row=row, column=col).value = policy.get(header, "")

    # Adjust column width for Route Policy tab
    for col in ws_route_policies.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_route_policies.column_dimensions[column].width = adjusted_width

    # Add filters to Route Policy tab headers
    last_col_letter_route_policy = openpyxl.utils.get_column_letter(len(route_policy_headers))
    ws_route_policies.auto_filter.ref = f"A1:{last_col_letter_route_policy}{len(route_policies) + 1}"

    ### Create IPSEC tab
    ws_ipsec = wb.create_sheet(title="IPSEC")

    # Create headers for IPSEC tab
    ipsec_headers = ["Source Name", "Interface", "KeyLife", "Peertype", "Algorithme", "Remote Adresse",
                     "Name Phase 2", "Réseaux Source Phase 2", "Réseaux Destination Phase 2"]

    for col, header in enumerate(ipsec_headers, 1):
        cell = ws_ipsec.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Fill IPSEC data with alternating row colors
    for row, ipsec in enumerate(ipsec_configs, 2):
        if row % 2 == 0:  # Even rows
            for col in range(1, len(ipsec_headers) + 1):
                ws_ipsec.cell(row=row, column=col).fill = light_fill

        # Fill data
        for col, header in enumerate(ipsec_headers, 1):
            ws_ipsec.cell(row=row, column=col).value = ipsec.get(header, "")

    # Adjust column width for IPSEC tab
    for col in ws_ipsec.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws_ipsec.column_dimensions[column].width = adjusted_width

    # Add filters to IPSEC tab headers
    last_col_letter_ipsec = openpyxl.utils.get_column_letter(len(ipsec_headers))
    ws_ipsec.auto_filter.ref = f"A1:{last_col_letter_ipsec}{len(ipsec_configs) + 1}"

    # Create filename with timestamp YYYYMMDD-HHMMSS-input.filename
    now = datetime.datetime.now()  # Current date based on your input
    timestamp = now.strftime("%Y%m%d-%H%M%S")
    input_filename = os.path.basename(input_file)
    output_filename = f"{timestamp}-{input_filename[:-5]}.xlsx"

    # Save Excel file
    wb.save(output_filename)
    return output_filename

def main():
    # Ask user for input file
    input_file = input("Veuillez entrer le chemin du fichier de configuration FortiGate: ")

    if not os.path.exists(input_file):
        print(f"Erreur: Le fichier '{input_file}' n'existe pas.")
        return

    try:
        # Parse the file
        print("Analyse du fichier en cours...")
        system_info = parse_system_info.parse_system_info(input_file)
        interfaces = parse_interfaces.parse_interfaces(input_file)
        dhcpv4_pools = parse_dhcp_pools.parse_dhcpv4_pools(input_file)
        dhcpv6_pools = parse_dhcp_pools.parse_dhcpv6_pools(input_file)
        hostsv4 = parse_hosts.parse_hostsv4(input_file)
        hostsv4 = parse_hosts.update_hostsv4_with_dhcp(input_file, hostsv4)
        hostsv4 = parse_hosts.update_hostsv4_with_ip_range(hostsv4)
        hostsv6 = parse_hosts.parse_hostsv6(input_file)
        address_groups = parse_address_groups.parse_address_groups(input_file)
        admins = parse_users.parse_admins(input_file)
        users = parse_users.parse_users(input_file)
        user_groups = parse_user_groups.parse_user_groups(input_file)
        zones = parse_zones.parse_zones(input_file, interfaces) 
        ip_pools = parse_ip_pools.parse_ip_pools(input_file)
        virtual_ips = parse_virtual_ips.parse_virtual_ips(input_file)
        config_firewall_policies = "config firewall policy"
        firewall_policies_v4 = parse_firewall_policies.parse_firewall_policies_v4(input_file,config_firewall_policies)
        config_firewall_policies = "config firewall policy6"
        firewall_policies_v6 = parse_firewall_policies.parse_firewall_policies_v4(input_file,config_firewall_policies)
        routes = parse_routes.parse_routes(input_file)
        route_policies = parse_route_policies.parse_route_policies(input_file)
        ipsec_phase1_configs = parse_ipsec.parse_ipsec_phase1(input_file)
        parse_ipsec.parse_ipsec_phase2(input_file, ipsec_phase1_configs)

        # Create Excel report
        print("Création du rapport Excel...")
        output_file = create_excel_report(system_info, interfaces, dhcpv4_pools, dhcpv6_pools, hostsv4, hostsv6, address_groups, admins, users, user_groups, zones, ip_pools, virtual_ips, firewall_policies_v4, firewall_policies_v6, routes, route_policies, ipsec_phase1_configs, input_file)

        print(f"Rapport Excel créé avec succès: {output_file}")

    except Exception as e:
        print(f"Une erreur s'est produite: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()